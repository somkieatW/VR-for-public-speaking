import pyaudio
import sys
import wave
import time
import zmq
import speech_recognition as sr
import threading
import voiceAnalysis as va
import os
import xlsxwriter as xl
import openpyxl as opxl
import requests
import json
import librosa
import datetime
import csv


exitFlag = 0
finish_record = False
count = 0

class myThread(threading.Thread):
    def __init__(self, threadID, name, counter):
        threading.Thread.__init__(self)
        self.threadID = threadID
        self.name = name
        self.counter = counter
    def run(self):
        if (self.counter == 1):
            signal_reciever()
        
        elif (self.counter == 2):
            record_all_the_time()
        
        elif (self.counter == 3):
            record_for_the_time(15)
        
        else:
            print("Error...")

def print_time(threadName, counter, delay):
   while counter:
        if (exitFlag):
            threadName.exit()
        time.sleep(delay)
        print ("%s: %s" % (threadName, time.ctime(time.time())))
        counter -= 1

def signal_reciever():
    context = zmq.Context()
    socket = context.socket(zmq.REP)
    socket.bind("tcp://*:5555")

    message = socket.recv()
    print("Recieved request: %s" %message)
    f= open("Result/guru99.txt","w+")
    f.write(message) 
    f.close()
    if (message == "Stop"):
        global finish_record
        finish_record = True
    socket.close()
def record_all_the_time():
    global count
    global RESULT_PATH
    global finish_record
    frames = []

    while finish_record == False:
        data = stream.read(CHUNK)
        frames.append(data)

    wf = wave.open(RESULT_PATH+".wav", 'wb')
    wf.setnchannels(CHANNELS)
    wf.setsampwidth(audio.get_sample_size(FORMAT))
    wf.setframerate(RATE)
    wf.writeframes(b''.join(frames))
    wf.close()

def record_for_the_time(time, name, c):
    global count
    global RESULT_PATH
    global finish_record
    frames = []
    
    for i in range(int(RATE / CHUNK * time)):
        if (finish_record == True):
            break
        # read audio stream from microphone
        data = stream.read(CHUNK)
        # append audio data to frames list
        frames.append(data)

    count = count+1
    wf = wave.open((RESULT_PATH+name+str(count)+".wav"), 'wb')
    wf.setnchannels(CHANNELS)
    wf.setsampwidth(audio.get_sample_size(FORMAT))
    wf.setframerate(RATE)
    wf.writeframes(b''.join(frames))
    wf.close()

def path_before_current():
    full_path = os.getcwd()
    split_path = full_path.split("\\")
    pbc = ""
    for folder_depth in range(len(split_path)):
        if (folder_depth == (len(split_path)-1)):
            break
        pbc += split_path[folder_depth]
        if (folder_depth < (len(split_path)-2)):
            pbc += "\\"
    return pbc            

def analyze_pause(file_name, path):
    x = va.mysppaus(file_name, path)
    return x

def wit(file_name, path):
    API_ENDPOINT = 'https://api.wit.ai/speech'
 
    # Wit.ai api access token
    wit_access_token = 'BBL4UDNORDRPGH5V5RDDHT5WRRQMWQQ7'

    audio = read_audio(path+"\\"+file_name+".wav")
 
    # defining headers for HTTP request
    headers = {'authorization': 'Bearer ' + wit_access_token,
               'Content-Type': 'audio/wav'}
 
    # making an HTTP post request
    resp = requests.post(API_ENDPOINT, headers = headers,
                         data = audio)
 
    # converting response content to JSON format
    data = json.loads(resp.content)
 
    # get text from data
    words = data['_text'].split(" ")
    amount_of_word = len(words)
 
    # return the text
    return amount_of_word, data['_text']

def read_audio(WAVE_FILENAME):
    # function to read audio(wav) file
    with open(WAVE_FILENAME, 'rb') as f:
        audio = f.read()
    return audio

def open_exel(workbook_path, workbook_name):
    try:
        workbook = opxl.load_workbook(workbook_path+"\\"+workbook_name+".xlsx")
    except IOError:
        workbook = opxl.Workbook()
        ws = workbook.active
        if (workbook_name=="Speech"):
            ws.append(["Username", "Times", "Amount of Word(s)", "Amount of Filler Word(s)", "Duration", "Date", "Time"])
        workbook.save(workbook_path+"\\"+workbook_name+".xlsx")
        workbook = opxl.load_workbook(workbook_path+"\\"+workbook_name+".xlsx")
    finally:
        return workbook

def add_to_column_in_last_row(workbook, workbook_name, username, amount_of_word, amount_of_filler_word, duration, date, time, workbook_path):
    worksheet = workbook.active
    last_row = worksheet.max_row
    times_in_the_last_row = 0
    for i in range(len(worksheet[1])):
        if worksheet[1][i+1].value == "Times":
            if(worksheet[last_row][i+1].value == "Times"):
                times_in_the_last_row = 0
            else:
                times_in_the_last_row = worksheet[last_row][i+1].value
            f= open("Result/half1.txt","w+")
            f.write("worksheet[1][i+1].value>>> " + str(times_in_the_last_row))
            f.close()
            break
    worksheet.append([username, (times_in_the_last_row+1), amount_of_word, amount_of_filler_word, duration, date, time])
    workbook.save(workbook_path + "\\" + workbook_name +".xlsx")
    
    f= open("Result/half2.txt","w+")
    f.write("str(amount_of_word)")
    f.close()

    wb = opxl.load_workbook(workbook_path + "\\" + workbook_name + '.xlsx')
    sh = wb.get_active_sheet()
    with open(workbook_path + "\\" + workbook_name + '.csv', 'wb') as f:  # open('test.csv', 'w', newline="") for python 3
        c = csv.writer(f)
        for r in sh.rows:
            c.writerow([cell.value for cell in r])

CHUNK = 1024
FORMAT = pyaudio.paInt16
CHANNELS = 2
RATE = 44100
RECORD_SECONDS = 10
RESULT_PATH = "Result"

audio = pyaudio.PyAudio()

stream = audio.open(format=FORMAT,
                channels=CHANNELS,
                rate=RATE,
                input=True,
                frames_per_buffer=CHUNK)

# Create new threads
thread1 = myThread(1, "Receiver", 1)
thread2 = myThread(2, "Main_Recorder", 2)
thread3 = myThread(3, "Sub_Recorder_1", 3)

name_of_audio_file = "output" # Audio File title
path_of_audio_file = os.getcwd() + "\\Result" # Path to the Audio_File directory (Python 3.7)

# Start new Threads
now = datetime.datetime.now().strftime("%H:%M:%S")

thread1.start()
thread2.start()
thread3.start()

thread1.join()
thread2.join()
thread3.join()

f= open("Result/finish.txt","w+")
f.close()

amount_of_filler_word = analyze_pause(name_of_audio_file, path_of_audio_file)
amount_of_word = wit(name_of_audio_file,path_of_audio_file)[0]
duration = int(librosa.get_duration(filename=(path_of_audio_file+"\\"+name_of_audio_file+'.wav')))
today = datetime.date.today()
workbook = open_exel(path_of_audio_file, "Speech")
add_to_column_in_last_row(workbook, "Speech", "Mr.None", amount_of_word, amount_of_filler_word, duration, today, now, path_of_audio_file)

f= open("Result/AllAreDone.txt","w+")
f.close()